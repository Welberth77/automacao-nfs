# main.py
import os
import importlib
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from certidoes import extrair_cnpj_do_pfx

PASTA_DADOS = "./dados"

SISTEMAS = {
    "sao_paulo":      "cidades.sao_paulo",
    "campinas":       "cidades.betha",
    "rio_de_janeiro": "cidades.rio_de_janeiro",
    "belo_horizonte": "cidades.betha",
}

COLUNAS_OBRIGATORIAS = [
    "CNPJ Prestador",
    "Certificado Digital .PFX",
    "Senha da Certidão",
    "CNPJ Tomador",
    "Valor",
]

def converter_pfx_se_necessario(caminho_pfx: str, senha: str) -> bool:
    from cryptography.hazmat.primitives.serialization import pkcs12
    from cryptography.hazmat.primitives.serialization.pkcs12 import serialize_key_and_certificates
    from cryptography.hazmat.primitives.serialization import BestAvailableEncryption
    from cryptography.hazmat.backends import default_backend

    with open(caminho_pfx, 'rb') as f:
        dados = f.read()

    try:
        pkcs12.load_key_and_certificates(dados, senha.encode(), default_backend())
        return True
    except Exception:
        pass

    try:
        from cryptography.hazmat.decrepit.ciphers.algorithms import RC2  # noqa
        chave, cert, extras = pkcs12.load_key_and_certificates(
            dados, senha.encode(), default_backend()
        )
        novo_pfx = serialize_key_and_certificates(
            name=None, key=chave, cert=cert, cas=extras,
            encryption_algorithm=BestAvailableEncryption(senha.encode())
        )
        with open(caminho_pfx, 'wb') as f:
            f.write(novo_pfx)
        print(f"  🔄 Certificado convertido: {os.path.basename(caminho_pfx)}")
        return True
    except Exception as e:
        print(f"  ⚠️ Não foi possível converter {os.path.basename(caminho_pfx)}: {e}")
        return False


def validar_planilha(df: pd.DataFrame) -> bool:
    print("\n🔎 Validando planilha...")
    erros = []

    for idx, row in df.iterrows():
        for col in COLUNAS_OBRIGATORIAS:
            valor = str(row.get(col, '')).strip()
            if not valor or valor == 'nan':
                erros.append(f"  ❌ Linha {idx + 2}: coluna '{col}' está vazia")

        cnpj = ''.join(filter(str.isdigit, str(row.get('CNPJ Prestador', ''))))
        if cnpj and len(cnpj) != 14:
            erros.append(f"  ❌ Linha {idx + 2}: CNPJ Prestador inválido → {row.get('CNPJ Prestador')}")

        cnpj_t = ''.join(filter(str.isdigit, str(row.get('CNPJ Tomador', ''))))
        if cnpj_t and len(cnpj_t) not in [11, 14]:
            erros.append(f"  ❌ Linha {idx + 2}: CNPJ Tomador inválido → {row.get('CNPJ Tomador')}")

        pfx = str(row.get('Certificado Digital .PFX', '')).strip()
        if pfx and pfx != 'nan' and not os.path.exists(pfx):
            erros.append(f"  ❌ Linha {idx + 2}: arquivo .pfx não encontrado → {pfx}")

    if erros:
        print(f"  ⚠️ {len(erros)} erro(s) encontrado(s):")
        for erro in erros:
            print(erro)
        print("\n  Corrija os erros acima antes de continuar.")
        return False

    print("  ✅ Planilha válida!")
    return True


def atualizar_status(planilha: str, linha_idx: int, status: str, numero_nota: str = "", observacao: str = ""):
    try:
        wb = load_workbook(planilha)
        ws = wb.active

        headers = {cell.value: cell.column for cell in ws[1]}
        row = linha_idx + 2

        if "Status" in headers:
            ws.cell(row=row, column=headers["Status"]).value = status
        if "Data de Emissão" in headers:
            ws.cell(row=row, column=headers["Data de Emissão"]).value = datetime.now().strftime("%d/%m/%Y %H:%M")
        if "NFS-e Emitida" in headers:
            ws.cell(row=row, column=headers["NFS-e Emitida"]).value = numero_nota
        if "Observação" in headers:
            ws.cell(row=row, column=headers["Observação"]).value = observacao

        wb.save(planilha)

    except PermissionError:
        print("\n" + "="*50)
        print("⚠️  ATENÇÃO: Não foi possível salvar a planilha!")
        print("   O arquivo está aberto no Excel.")
        print("   Feche o arquivo e rode novamente.")
        print("="*50 + "\n")


def carregar_certificados_da_cidade(cidade: str, df: pd.DataFrame) -> dict:
    pasta_certs = os.path.join(PASTA_DADOS, cidade, "certidoes")
    certificados = {}

    if not os.path.exists(pasta_certs):
        print(f"⚠️ Pasta de certidões não encontrada: {pasta_certs}")
        return certificados

    senhas = {}
    for _, row in df.iterrows():
        pfx = str(row.get('Certificado Digital .PFX', '')).strip()
        senha = str(row.get('Senha da Certidão', '')).strip()
        if pfx and pfx != 'nan':
            senhas[os.path.basename(pfx)] = senha

    for arquivo in os.listdir(pasta_certs):
        if not arquivo.endswith(".pfx"):
            continue

        caminho_pfx = os.path.join(pasta_certs, arquivo)
        senha = senhas.get(arquivo, "")

        if not senha:
            print(f"  ⚠️ Senha não encontrada para {arquivo}, pulando...")
            continue

        converter_pfx_se_necessario(caminho_pfx, senha)

        try:
            cnpj = extrair_cnpj_do_pfx(caminho_pfx, senha)
            certificados[cnpj] = {"path": caminho_pfx, "password": senha, "cnpj": cnpj}
            print(f"  ✅ {arquivo} → CNPJ: {cnpj}")
        except Exception as e:
            print(f"  ⚠️ Erro ao carregar {arquivo}: {e}")

    return certificados


def main():
    for cidade in os.listdir(PASTA_DADOS):
        caminho_cidade = os.path.join(PASTA_DADOS, cidade)

        if not os.path.isdir(caminho_cidade):
            continue

        if cidade not in SISTEMAS:
            print(f"⚠️ Cidade '{cidade}' não tem sistema mapeado, pulando...")
            continue

        print(f"\n🏙️ Processando cidade: {cidade.upper()}")

        planilha = os.path.join(caminho_cidade, f"{cidade}.xlsx")
        if not os.path.exists(planilha):
            print(f"⚠️ Planilha não encontrada: {planilha}, pulando...")
            continue

        df = pd.read_excel(planilha, dtype=str)
        df.columns = df.columns.str.strip()
        df = df[df["CNPJ Prestador"].notna()]
        df = df[df["CNPJ Prestador"] != "12.345.678/0001-99"]

        if df.empty:
            print(f"⚠️ Nenhuma nota encontrada em {cidade}, pulando...")
            continue

        if not validar_planilha(df):
            print(f"⚠️ Pulando {cidade} por erros na planilha.")
            continue

        print("🔍 Carregando certidões...")
        certificados = carregar_certificados_da_cidade(cidade, df)

        if not certificados:
            print(f"⚠️ Nenhuma certidão encontrada para {cidade}, pulando...")
            continue

        # Filtra apenas PENDENTES
        if "Status" in df.columns:
            df_pendentes = df[df["Status"].str.upper() == "PENDENTE"].copy()
        else:
            df_pendentes = df.copy()

        modulo = importlib.import_module(SISTEMAS[cidade])

        print(f"📋 {len(df_pendentes)} nota(s) pendente(s) em {cidade}")

        for idx, (original_idx, nota) in enumerate(df_pendentes.iterrows()):
            cnpj_prestador = ''.join(filter(str.isdigit, nota['CNPJ Prestador']))
            cert = certificados.get(cnpj_prestador)

            if not cert:
                print(f"⚠️ Certidão não encontrada para CNPJ: {cnpj_prestador}, pulando...")
                atualizar_status(planilha, original_idx, "ERRO", observacao="Certidão não encontrada")
                continue

            modulo.emitir_nfse(cert, nota.to_dict(), original_idx, planilha)

    print("\n🎉 Processo finalizado para todas as cidades!")

if __name__ == "__main__":
    main()