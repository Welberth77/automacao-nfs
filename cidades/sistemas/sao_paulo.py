import json
import time
import pandas as pd
from playwright.sync_api import sync_playwright
from certificados import extrair_cnpj_do_pfx

URL_NFSE_SP = "https://nfe.prefeitura.sp.gov.br/login.aspx"
PLANILHA = "./nfse_emissao.xlsx"

def aguardar(page, segundos=2):
    """Espera a página carregar completamente."""
    page.wait_for_load_state("networkidle")
    time.sleep(segundos)

def carregar_certificados():
    with open('./config.json', 'r') as f:
        config = json.load(f)

    certificados = []
    for cert in config['certificados']:
        try:
            cnpj = extrair_cnpj_do_pfx(cert['path'], cert['password'])
            certificados.append({
                "path": cert['path'],
                "password": cert['password'],
                "cnpj": cnpj
            })
            print(f"✅ Carregado: {cert['path']} → CNPJ: {cnpj}")
        except Exception as e:
            print(f"⚠️ Erro ao carregar {cert['path']}: {e}")

    return certificados

def carregar_notas():
    """Lê a planilha e retorna apenas as notas com status PENDENTE."""
    df = pd.read_excel(PLANILHA, sheet_name="Notas", dtype=str)
    df.columns = df.columns.str.strip()
    pendentes = df[df["STATUS"].str.upper() == "PENDENTE"].copy()
    print(f"📋 {len(pendentes)} nota(s) pendente(s) encontrada(s) na planilha")
    return pendentes

def atualizar_status(linha_idx: int, status: str, observacao: str = ""):
    """Atualiza o status da linha na planilha após emissão."""
    from openpyxl import load_workbook
    from datetime import datetime

    wb = load_workbook(PLANILHA)
    ws = wb["Notas"]

    row = linha_idx + 3
    ws.cell(row=row, column=10).value = status
    ws.cell(row=row, column=11).value = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws.cell(row=row, column=13).value = observacao
    wb.save(PLANILHA)


def emitir_nfse(cert: dict, nota: dict, linha_idx: int):
    print(f"\n🚀 Emitindo nota para tomador: {nota['TOMADOR']} | CNPJ: {nota['CNPJ TOMADOR']}")

    with open(cert['path'], 'rb') as f:
        pfx_bytes = f.read()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            ignore_https_errors=True,
            client_certificates=[
                {
                    "origin": "https://nfe.prefeitura.sp.gov.br",
                    "pfx": pfx_bytes,
                    "passphrase": cert['password']
                }
            ]
        )

        page = context.new_page()

        try:
            # Passo 1 — Acessa o site
            print("🌐 Acessando NFS-e SP...")
            page.goto(URL_NFSE_SP)
            aguardar(page, 2)

            # Passo 2 — Clica em Login único
            print("🖱️ Clicando em Login único...")
            page.click("button.oauth-button")
            aguardar(page, 2)
            print("✅ Clicou em Login único!")

            # Passo 3 — Via Certificado ICP-Brasil
            print("🖱️ Clicando em Via Certificado ICP-Brasil...")
            page.click("#btnCertificado")
            aguardar(page, 3)
            print("✅ Certificado enviado automaticamente!")

            # Passo 4 — Acessar o Sistema
            print("🖱️ Clicando em Acessar o Sistema...")
            page.wait_for_selector("#ctl00_body_btAcesso", timeout=10000)
            page.click("#ctl00_body_btAcesso")
            aguardar(page, 3)
            print("✅ Acessou o sistema!")

            # Passo 5 — Emissão de NFS-e
            print("🖱️ Clicando em Emissão de NFS-e...")
            page.wait_for_selector("a[href='nota.aspx']", timeout=10000)
            page.click("a[href='nota.aspx']")
            aguardar(page, 2)
            print("✅ Na página de emissão de NFS-e!")

            # Passo 6 — Lê o CNPJ do topo da página
            print("🔍 Lendo CNPJ do topo da página...")
            page.wait_for_selector("div.oauth-name", timeout=10000)
            cnpj_topo = page.inner_text("div.oauth-name")
            cnpj_tomador = cnpj_topo.strip()
            print(f"✅ CNPJ lido: {cnpj_tomador}")

            # Passo 7 — Preenche CNPJ do tomador
            print("✏️ Preenchendo CNPJ do tomador...")
            page.wait_for_selector("#ctl00_body_tbCPFCNPJTomador", timeout=10000)
            page.fill("#ctl00_body_tbCPFCNPJTomador", cnpj_tomador)
            aguardar(page, 1)
            page.click("#ctl00_body_btAvancar")
            aguardar(page, 3)
            print("✅ CNPJ do tomador preenchido!")

            # Passo 8 — Lê o código pré-selecionado no dropdown
            print("🔍 Lendo código pré-selecionado no dropdown...")
            frame = page
            for f in page.frames:
                if f.query_selector("#ctl00_body_ddlAtividade"):
                    frame = f
                    print(f"  → Frame encontrado: {f.url}")
                    break

            frame.wait_for_selector("#ctl00_body_ddlAtividade", timeout=10000)
            codigo_selecionado = frame.eval_on_selector(
                "#ctl00_body_ddlAtividade",
                "select => select.options[select.selectedIndex].value"
            )
            print(f"✅ Código pré-selecionado: {codigo_selecionado}")

            # Passo 9 — Digita o código e clica em >>
            print(f"✏️ Digitando código do serviço: {codigo_selecionado}...")
            frame.wait_for_selector("#ctl00_body_tbServEncerradoCodigo", timeout=10000)
            frame.fill("#ctl00_body_tbServEncerradoCodigo", codigo_selecionado)
            aguardar(frame, 1)
            frame.click("#ctl00_body_btnServEncerradoSelecionar")
            aguardar(frame, 3)
            print("✅ Código do serviço selecionado!")

            # Passo 10 — Preenche descrição do serviço
            print("✏️ Preenchendo descrição do serviço...")
            frame.wait_for_selector("#ctl00_body_tbDiscriminacao", timeout=10000)
            frame.fill("#ctl00_body_tbDiscriminacao", "1 Serviço Prestado: ")
            aguardar(frame, 1)
            print("✅ Descrição preenchida!")

            # Passo 11 — Preenche o valor do serviço (vem da planilha)
            print(f"✏️ Preenchendo valor: {nota['VALOR (R$)']}...")
            frame.wait_for_selector("#ctl00_body_tbValor", timeout=10000)
            frame.fill("#ctl00_body_tbValor", str(nota['VALOR (R$)']).strip())
            aguardar(frame, 1)
            print("✅ Valor preenchido!")

            # ⚠️ TODO — DESCOMENTAR QUANDO FOR FINALIZAR O PROJETO
            # Passo 12 — Clica em EMITIR
            # print("🖱️ Clicando em EMITIR...")
            # frame.wait_for_selector("#ctl00_body_btEmitir", timeout=10000)
            # frame.click("#ctl00_body_btEmitir")
            # aguardar(frame, 3)
            # print("✅ Nota fiscal emitida!")

            atualizar_status(linha_idx, "OK")
            print("✅ Formulário preenchido com sucesso!")

        except Exception as e:
            print(f"❌ Erro: {e}")
            atualizar_status(linha_idx, "ERRO", str(e))

        finally:
            input("⏸️ Pressione Enter para continuar para a próxima nota...")
            context.close()  # ← desvincula o certificado da sessão
            browser.close()  # ← fecha o browser


if __name__ == "__main__":
    print("🔍 Carregando certificados...\n")
    certificados = carregar_certificados()
    certificados_por_cnpj = {c['cnpj']: c for c in certificados}

    print(f"\n📋 {len(certificados)} certificado(s) carregado(s)")

    notas = carregar_notas()

    for idx, (_, nota) in enumerate(notas.iterrows()):
        cnpj_prestador = ''.join(filter(str.isdigit, nota['CNPJ PRESTADOR']))
        cert = certificados_por_cnpj.get(cnpj_prestador)

        if not cert:
            print(f"⚠️ Certificado não encontrado para CNPJ: {cnpj_prestador}, pulando...")
            atualizar_status(idx, "ERRO", "Certificado .pfx não encontrado")
            continue

        emitir_nfse(cert, nota.to_dict(), idx)

    print("\n🎉 Processo finalizado para todas as notas!")